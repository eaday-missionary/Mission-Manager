"""Excel import and normalization helpers."""

from __future__ import annotations

from datetime import datetime, time
from pathlib import Path
from typing import Any, Iterable

from .constants import BOOLEAN_FALSE, BOOLEAN_TRUE, CANONICAL_HEADERS, HEADER_ALIASES, HEADER_TO_FIELD, PERSON_FIELDS
from .models import ParsedDataset, ValidationError


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def canonicalize_header(header: Any) -> str:
    raw = str(header).strip() if header is not None else ""
    return HEADER_ALIASES.get(raw, raw)


def normalize_boolean(value: Any) -> tuple[bool | None, str | None]:
    text = normalize_text(value)
    if text is None:
        return None, None
    lowered = text.lower()
    if lowered in BOOLEAN_TRUE:
        return True, None
    if lowered in BOOLEAN_FALSE:
        return False, None
    return None, f"Unknown boolean literal '{text}' normalized to null"


def _parse_time_text(text: str) -> str | None:
    for fmt in ("%H:%M", "%H%M", "%I:%M %p"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.strftime("%H:%M")
        except ValueError:
            continue
    return None


def normalize_time(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, (int, float)):
        frac = float(value) % 1.0
        total_minutes = int(round(frac * 24 * 60)) % (24 * 60)
        hh = total_minutes // 60
        mm = total_minutes % 60
        return f"{hh:02d}:{mm:02d}"
    text = normalize_text(value)
    if text is None:
        return None
    return _parse_time_text(text)


def validate_headers(headers: list[str]) -> tuple[list[str], list[str]]:
    found = {canonicalize_header(h) for h in headers if h is not None and str(h).strip()}
    required = set(CANONICAL_HEADERS)
    missing = sorted(required - found)
    unexpected = sorted(h for h in found - required if h)
    return missing, unexpected


def _identity_key(record: dict[str, Any]) -> tuple[str, str, str]:
    return (
        (record.get("first_name") or "").lower(),
        (record.get("last_name") or "").lower(),
        (record.get("current_area") or "").lower(),
    )


def parse_records_from_rows(headers: list[str], rows: Iterable[Iterable[Any]], source_file_name: str, row_start: int = 2) -> ParsedDataset:
    missing, unexpected = validate_headers(headers)
    errors: list[ValidationError] = []
    warnings: list[str] = []
    if missing:
        errors.append(
            ValidationError(
                code="SCHEMA_ERROR",
                message="Missing required headers",
                suggested_action=f"Add headers: {', '.join(missing)}",
            )
        )
    if errors:
        return ParsedDataset([], 0, 0, errors, warnings)

    index_for: dict[str, int] = {}
    for idx, header in enumerate(headers):
        canonical = canonicalize_header(header)
        if canonical in HEADER_TO_FIELD:
            index_for[canonical] = idx
    deduped: dict[tuple[str, str, str], dict[str, Any]] = {}
    processed = 0
    skipped = 0

    for row_number, row_values in enumerate(rows, start=row_start):
        row_list = list(row_values)
        if not any(normalize_text(v) is not None for v in row_list):
            skipped += 1
            continue

        processed += 1
        record: dict[str, Any] = {f: None for f in PERSON_FIELDS}
        record["source_file_name"] = source_file_name
        record["source_row_number"] = row_number

        for header, idx in index_for.items():
            field = HEADER_TO_FIELD[header]
            raw = row_list[idx] if idx < len(row_list) else None
            if field in ("staying", "second_leg"):
                bval, warn = normalize_boolean(raw)
                record[field] = bval
                if warn:
                    warnings.append(f"Row {row_number} {field}: {warn}")
            elif field in ("departure_time", "arrival_time", "second_departure_time", "second_arrival_time"):
                tval = normalize_time(raw)
                if normalize_text(raw) is not None and tval is None:
                    errors.append(
                        ValidationError(
                            code="ROW_VALIDATION_ERROR",
                            message=f"Invalid time value '{raw}'",
                            field=field,
                            row_number=row_number,
                            suggested_action="Use HH:mm format or valid Excel time value",
                        )
                    )
                record[field] = tval
            else:
                record[field] = normalize_text(raw)

        if not record["first_name"] or not record["last_name"]:
            errors.append(
                ValidationError(
                    code="ROW_VALIDATION_ERROR",
                    message="Missing required identity field",
                    row_number=row_number,
                    suggested_action="Provide First Name and Last Name",
                )
            )
            continue

        deduped[_identity_key(record)] = record

    if unexpected:
        warnings.append(f"Ignoring unknown columns: {', '.join(unexpected)}")

    return ParsedDataset(list(deduped.values()), processed, skipped, errors, warnings)


def parse_excel_file(file_path: str) -> ParsedDataset:
    path = Path(file_path)
    if not path.exists() or not path.is_file():
        return ParsedDataset([], 0, 0, [ValidationError(code="FILE_ERROR", message="File not found")], [])

    ext = path.suffix.lower()
    if ext in {".xlsx", ".xlsm"}:
        return _parse_openpyxl(path)
    if ext == ".xls":
        return _parse_xls(path)

    return ParsedDataset(
        [],
        0,
        0,
        [ValidationError(code="FILE_ERROR", message="Unsupported file type", suggested_action="Use .xlsx, .xlsm, or .xls")],
        [],
    )


def _parse_openpyxl(path: Path) -> ParsedDataset:
    try:
        import openpyxl
    except ImportError:
        return ParsedDataset([], 0, 0, [ValidationError(code="FILE_ERROR", message="openpyxl is not installed")], [])

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.worksheets[0]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            return ParsedDataset([], 0, 0, [ValidationError(code="SCHEMA_ERROR", message="Header row is missing")], [])
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        rows = ws.iter_rows(min_row=2, values_only=True)
        return parse_records_from_rows(headers, rows, path.name, row_start=2)
    except Exception as exc:
        return ParsedDataset([], 0, 0, [ValidationError(code="FILE_ERROR", message=f"Failed reading workbook: {exc}")], [])


def _parse_xls(path: Path) -> ParsedDataset:
    try:
        import xlrd
    except ImportError:
        return ParsedDataset([], 0, 0, [ValidationError(code="FILE_ERROR", message="xlrd is not installed")], [])

    try:
        wb = xlrd.open_workbook(path)
        sheet = wb.sheet_by_index(0)
        if sheet.nrows < 1:
            return ParsedDataset([], 0, 0, [ValidationError(code="SCHEMA_ERROR", message="Header row is missing")], [])

        headers = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]

        def iter_rows() -> Iterable[list[Any]]:
            for r in range(1, sheet.nrows):
                yield [sheet.cell_value(r, c) for c in range(sheet.ncols)]

        return parse_records_from_rows(headers, iter_rows(), path.name, row_start=2)
    except Exception as exc:
        return ParsedDataset([], 0, 0, [ValidationError(code="FILE_ERROR", message=f"Failed reading .xls workbook: {exc}")], [])
